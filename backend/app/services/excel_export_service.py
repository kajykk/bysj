from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from typing import Any

logger = logging.getLogger(__name__)


@dataclass
class ExcelExportResult:
    """Result of Excel export."""

    success: bool
    excel_bytes: bytes | None = None
    file_size: int = 0
    row_count: int = 0
    column_count: int = 0
    error_message: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "success": self.success,
            "file_size": self.file_size,
            "row_count": self.row_count,
            "column_count": self.column_count,
            "error_message": self.error_message,
        }


@dataclass
class ExcelStreamResult:
    """RES-P2-003: Excel 流式导出结果.

    与 ExcelExportResult 区别: 保留 BytesIO 流 (不调用 getvalue()),
    避免 Excel 完整 bytes 在内存中再拷贝一份. API 层用 StreamingResponse
    分块读取 BytesIO 实现真正的流式响应.

    适用于: 大数据量 Excel 导出 (> 5000 行), 避免内存峰值.
    """

    success: bool
    stream: io.BytesIO | None = None
    file_size: int = 0
    row_count: int = 0
    column_count: int = 0
    error_message: str | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "success": self.success,
            "file_size": self.file_size,
            "row_count": self.row_count,
            "column_count": self.column_count,
            "error_message": self.error_message,
        }


class ExcelExportService:
    """Export data to Excel (.xlsx) using openpyxl.

    Features:
    - Support 10000+ rows export
    - Column filtering and filtering conditions
    - Streaming write to prevent memory overflow
    - RES-P2-003: 流式输出 (export_to_stream) 返回 BytesIO, 避免 bytes 拷贝
    """

    BATCH_SIZE = 1000
    # RES-P2-003: 流式输出阈值, 超过此行数自动切换到 export_to_stream
    STREAM_THRESHOLD_ROWS = 5000

    def __init__(self) -> None:
        pass

    def export_to_excel(
        self, data: list[dict[str, Any]], filename: str | None = None
    ) -> ExcelExportResult:
        """Backward-compatible Excel export wrapper used by performance tests."""
        return self.export(data=data)

    def _check_openpyxl_available(self) -> bool:
        """Check if openpyxl is installed."""
        try:
            import openpyxl  # noqa: F401

            return True
        except ImportError:
            return False

    def export(
        self,
        data: list[dict[str, Any]],
        columns: list[str] | None = None,
        filters: dict[str, Any] | None = None,
    ) -> ExcelExportResult:
        """Export data to Excel.

        Args:
            data: List of dictionaries with data.
            columns: List of columns to include. If None, use all keys from first row.
            filters: Dictionary of column -> value to filter by.

        Returns:
            ExcelExportResult with generated Excel file or error.
        """
        if not self._check_openpyxl_available():
            return ExcelExportResult(
                success=False,
                error_message="openpyxl not installed",
            )

        try:
            from openpyxl import Workbook

            # L-09 修复：移除未使用的 Font 导入

            if not data:
                return ExcelExportResult(
                    success=False,
                    error_message="No data to export",
                )

            # Determine columns
            if columns is None:
                columns = list(data[0].keys())

            # Apply filters
            filtered_data = self._apply_filters(data, filters)

            # Create workbook
            wb = Workbook(write_only=True)
            ws = wb.create_sheet(title="Data")

            # Write header
            ws.append(columns)

            # Write data in batches
            row_count = 0
            for i in range(0, len(filtered_data), self.BATCH_SIZE):
                batch = filtered_data[i : i + self.BATCH_SIZE]
                for row in batch:
                    row_values = [self._format_value(row.get(col)) for col in columns]
                    ws.append(row_values)
                    row_count += 1

            # Save to buffer
            buffer = io.BytesIO()
            try:
                wb.save(buffer)
                excel_bytes = buffer.getvalue()
            finally:
                buffer.close()

            return ExcelExportResult(
                success=True,
                excel_bytes=excel_bytes,
                file_size=len(excel_bytes),
                row_count=row_count,
                column_count=len(columns),
            )

        except Exception as exc:
            logger.exception("Excel export failed")
            return ExcelExportResult(
                success=False,
                error_message=str(exc),
            )

    def export_to_stream(
        self,
        data: list[dict[str, Any]],
        columns: list[str] | None = None,
        filters: dict[str, Any] | None = None,
    ) -> ExcelStreamResult:
        """RES-P2-003: Export data to Excel and return BytesIO stream.

        与 export 区别: 返回 BytesIO 流 (不调用 getvalue()), 避免 bytes 拷贝.
        适用于 StreamingResponse 流式响应.

        内部仍用 openpyxl Workbook(write_only=True) 流式写入 (已有优化),
        但避免最后 getvalue() 的内存拷贝 (大 Excel 可节省一倍内存).

        Args:
            data: List of dictionaries with data.
            columns: List of columns to include. If None, use all keys from first row.
            filters: Dictionary of column -> value to filter by.

        Returns:
            ExcelStreamResult with BytesIO stream (位置指针已 rewind 到 0).
        """
        if not self._check_openpyxl_available():
            return ExcelStreamResult(
                success=False,
                error_message="openpyxl not installed",
            )

        try:
            from openpyxl import Workbook

            if not data:
                return ExcelStreamResult(
                    success=False,
                    error_message="No data to export",
                )

            # Determine columns
            if columns is None:
                columns = list(data[0].keys())

            # Apply filters
            filtered_data = self._apply_filters(data, filters)

            # Create workbook (write_only=True 已是流式写入模式)
            wb = Workbook(write_only=True)
            ws = wb.create_sheet(title="Data")

            # Write header
            ws.append(columns)

            # Write data in batches
            row_count = 0
            for i in range(0, len(filtered_data), self.BATCH_SIZE):
                batch = filtered_data[i : i + self.BATCH_SIZE]
                for row in batch:
                    row_values = [self._format_value(row.get(col)) for col in columns]
                    ws.append(row_values)
                    row_count += 1

            # RES-P2-003: 保留 BytesIO 不 close, 由调用方 (StreamingResponse) 负责
            buffer = io.BytesIO()
            wb.save(buffer)
            # 不调用 getvalue(), 直接获取 size 后 rewind 指针
            file_size = buffer.tell()
            buffer.seek(0)

            return ExcelStreamResult(
                success=True,
                stream=buffer,
                file_size=file_size,
                row_count=row_count,
                column_count=len(columns),
            )

        except Exception as exc:
            logger.exception("Excel stream export failed")
            return ExcelStreamResult(
                success=False,
                error_message=str(exc),
            )

    def should_use_stream(self, data_row_count: int) -> bool:
        """RES-P2-003: 判断是否应使用流式输出.

        Args:
            data_row_count: 数据行数.

        Returns:
            True if stream output is recommended (> STREAM_THRESHOLD_ROWS).
        """
        return data_row_count >= self.STREAM_THRESHOLD_ROWS

    def _apply_filters(
        self,
        data: list[dict[str, Any]],
        filters: dict[str, Any] | None,
    ) -> list[dict[str, Any]]:
        """Apply filters to data.

        Args:
            data: Original data.
            filters: Dictionary of column -> value to filter by.

        Returns:
            Filtered data.
        """
        if not filters:
            return data

        filtered = []
        for row in data:
            match = True
            for col, value in filters.items():
                if row.get(col) != value:
                    match = False
                    break
            if match:
                filtered.append(row)

        return filtered

    def _format_value(self, value: Any) -> Any:
        """Format a value for Excel output.

        FE-003 修复：对以 =, +, -, @, \\t, \\r 开头的字符串进行转义，
        防止 Excel 公式注入攻击（CSV/Formula Injection）。
        M-Svc-2 修复：数字类型保持原样不转字符串，避免 Excel 将数值当作文本处理
        （数值列保留为数字类型后，Excel 可正确排序、求和、应用公式）。

        Args:
            value: Value to format.

        Returns:
            Formatted and sanitized value（数字保持原类型，其余为字符串）。
        """
        if value is None:
            return ""
        if isinstance(value, bool):
            return "Yes" if value else "No"
        # M-Svc-2 修复：int/float 保持原样不转字符串（bool 已在上一步处理）
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, (list, dict)):
            import json

            return self._sanitize_cell_value(json.dumps(value, ensure_ascii=False))
        return self._sanitize_cell_value(str(value))

    @staticmethod
    def _sanitize_cell_value(value: str) -> str:
        """FE-003 修复：转义 Excel 公式注入危险字符。

        当单元格值以 ``=``、``+``、``-``、``@``、``\\t``、``\\r`` 开头时，
        Excel 会将其解释为公式，可能导致公式注入攻击（如 ``=HYPERLINK(...)``、
        ``=cmd|'/c calc'!A1``）。通过在前面添加单引号 ``'`` 强制 Excel 将其
        作为文本处理（单引号在 Excel 中不会显示）。

        对于 ``-`` 开头的值，仅当其不是合法数字时才转义（避免影响负数导出）。
        """
        if not value:
            return value
        # 危险前缀字符
        dangerous_prefixes = ("=", "+", "@", "\t", "\r")
        if value.startswith(dangerous_prefixes):
            return "'" + value
        # 对 "-" 开头的值：仅当不是合法数字时才转义
        if value.startswith("-"):
            try:
                float(value)
            except ValueError:
                return "'" + value
        return value

    def export_large_dataset(
        self,
        data_generator,
        columns: list[str],
        filters: dict[str, Any] | None = None,
    ) -> ExcelExportResult:
        """Export large dataset using generator for memory efficiency.

        Args:
            data_generator: Generator yielding dictionaries.
            columns: List of columns to include.
            filters: Dictionary of column -> value to filter by.

        Returns:
            ExcelExportResult.
        """
        if not self._check_openpyxl_available():
            return ExcelExportResult(
                success=False,
                error_message="openpyxl not installed",
            )

        try:
            from openpyxl import Workbook

            wb = Workbook(write_only=True)
            ws = wb.create_sheet(title="Data")

            # Write header
            ws.append(columns)

            row_count = 0
            for row in data_generator:
                if filters and not self._row_matches_filters(row, filters):
                    continue

                row_values = [self._format_value(row.get(col)) for col in columns]
                ws.append(row_values)
                row_count += 1

            buffer = io.BytesIO()
            try:
                wb.save(buffer)
                excel_bytes = buffer.getvalue()
            finally:
                buffer.close()

            return ExcelExportResult(
                success=True,
                excel_bytes=excel_bytes,
                file_size=len(excel_bytes),
                row_count=row_count,
                column_count=len(columns),
            )

        except Exception as exc:
            logger.exception("Large Excel export failed")
            return ExcelExportResult(
                success=False,
                error_message=str(exc),
            )

    def _row_matches_filters(
        self, row: dict[str, Any], filters: dict[str, Any]
    ) -> bool:
        """Check if a row matches all filters.

        Args:
            row: Data row.
            filters: Filters to apply.

        Returns:
            True if row matches all filters.
        """
        for col, value in filters.items():
            if row.get(col) != value:
                return False
        return True


# Global service instance
excel_export_service = ExcelExportService()
